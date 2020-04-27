import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('pythontest.xlsx')
sheet = wb['Sheet1']

# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
    print()

values = Reference(sheet, min_row=2, max_row=sheet.max_row,min_col=4,max_col=4)

chart =BarChart()
chart.add_data(values)
sheet.add_chart(chart ,'e2')
wb.save('pythontest.xlsx')
